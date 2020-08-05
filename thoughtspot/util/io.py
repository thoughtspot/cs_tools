"""
Copyright 2020 ThoughtSpot
Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation the
rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to
permit persons to whom the Software is furnished to do so, subject to the following conditions:
The above copyright notice and this permission notice shall be included in all copies or substantial portions
of the Software.
THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED
TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
"""

from openpyxl import Workbook
from rich.console import Console
import rich.table as rt

from .model import Table, DependencyTree


class DependencyTreeStdoutWriter:
    """
    Writes a dependency tree to standard out.
    """

    def __init__(self):
        """Creates a new writer."""
        pass

    @staticmethod
    def write_dependency_tree(dt, rich_print=True):
        """
        Writes the dependency tree to standard out.
        :param dt: The dependency tree to write.
        :type dt: DependencyTree
        :param rich_print: If true, writes using tables and fonts.
        :param rich_print: bool
        :return: Nothing
        """
        if rich_print:
            DependencyTreeStdoutWriter.__write_rich_print(dt)
        else:
            root_tables = dt.get_root_tables()
            for table in root_tables:
                DependencyTreeStdoutWriter.__write_table_details(table=table, depth=0)
                DependencyTreeStdoutWriter.__write_dependents(dt=dt, table=table, depth=1)

    @staticmethod
    def __write_rich_print(dt):
        """
        Writes the dependency tree as a table to stdout.
        :param dt: The dependency tree to write.
        :type dt: DependencyTree
        :return: Nothing
        """

        console = Console()
        header = ["Name", "GUID", "Type", "Author", "Depends On", "Dependents"]
        table = rt.Table(show_header=True, style="bold", show_lines=True)
        for h in header:
            table.add_column(h)

        for t in dt.get_all_tables():
            dependents = dt.get_names_for_ids(dt.get_dependents(table_guid=t.id))
            dependents = [d if d else "" for d in dependents]  # Make sure not None
            dependents = "" if not dependents else '\n'.join(dependents)

            depends_on = dt.get_names_for_ids(dt.get_depends_on(table_guid=t.id))
            depends_on = [d if d else "" for d in depends_on]  # Make sure not None
            depends_on = "" if not depends_on else '\n'.join(depends_on)

            table.add_row(
                t.get_long_name(),
                t.id,
                t.type,
                t.authorName,
                depends_on,
                dependents,
            )

        console.print(table)

    @staticmethod
    def __write_dependents(dt, table, depth):
        """
        Writes the dependents of the given table to standard out at the depth (indenting two spaces).  This call
        is recursive so that all dependencies are under the parent.  It is assumed that there are no circular
        dependencies.
        :param dt: The dependency tree being written.
        :type dt: DependencyTree
        :param table: The table to write dependents for.
        :type table: Table
        :return: Nothing
        """
        dependencies = dt.get_dependents(table_guid=table.id)
        for dependent in dependencies:
            dependent_table = dt.get_table(dependent)
            DependencyTreeStdoutWriter.__write_table_details(dependent_table, depth=depth)
            DependencyTreeStdoutWriter.__write_dependents(dt=dt, table=dependent_table, depth=depth+1)

    @staticmethod
    def __write_table_details(table, depth):
        """
        Writes the table details, indented based on the depth.
        :param table: The table to write.
        :type table: Table
        :param depth: How far to indent.
        :type depth: int
        :return: Nothing
        """
        indent_size = 2  # number of spaces for indenting.

        table_str = f"{table.get_long_name()} ({table.id}): [type='{table.type}', " \
                    f"author='{table.authorDisplayName}]' "
        print(f'{" "*depth*indent_size}{">"*depth} {table_str}')


class DependencyTreeXLSWriter:
    """
    Writes a dependency tree to Excel.
    """

    def __init__(self):
        """
        Creates a new writer.
        """
        pass

    @staticmethod
    def write_to_excel(dt, filename):
        """
        Writes the dependency tree to the given file.
        :param dt: The dependency tree to write.
        :type dt: DependencyTree
        :param filename: The file to write to. .xlsx will be appended if it's not on the name.
        :type filename: str
        :return: Nothing
        """
        workbook = Workbook()
        workbook.remove(workbook.active)

        ws = workbook.create_sheet(title="Dependencies")
        header = ["Name", "GUID", "Type", "Author", "Depends On", "Dependents"]
        for ccnt in range(0, len(header)):
            ws.cell(column=ccnt+1, row=1, value=header[ccnt])

        rcnt = 2
        for table in dt.get_all_tables():
            dependents = dt.get_names_for_ids(dt.get_dependents(table_guid=table.id))
            depends_on = dt.get_names_for_ids(dt.get_depends_on(table_guid=table.id))

            ws.cell(column=1, row=rcnt, value=table.get_long_name())
            ws.cell(column=2, row=rcnt, value=table.id)
            ws.cell(column=3, row=rcnt, value=table.type)
            ws.cell(column=4, row=rcnt, value=table.authorName)
            ws.cell(column=5, row=rcnt, value=f"[{', '.join(depends_on)}]")
            ws.cell(column=6, row=rcnt, value=f"[{', '.join(dependents)}]")

            rcnt += 1

        if not filename.endswith('xlsx'):
            filename += ".xlsx"
        workbook.save(filename=filename)
