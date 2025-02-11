from __future__ import annotations

from typing import Literal, Optional, Union
import datetime as dt
import logging
import pathlib

import openpyxl
import pydantic

from cs_tools import _types
from cs_tools.sync import utils as sync_utils
from cs_tools.sync.base import Syncer

log = logging.getLogger(__name__)


class Excel(Syncer):
    """Interact with an Excel workbook."""

    __manifest_path__ = pathlib.Path(__file__).parent / "MANIFEST.json"
    __syncer_name__ = "excel"

    filepath: Union[pydantic.FilePath, pydantic.NewPath]
    filepath_suffix: Optional[str] = None
    date_time_format: str = sync_utils.DATETIME_FORMAT_ISO_8601
    save_strategy: Literal["APPEND", "OVERWRITE"] = "OVERWRITE"

    @pydantic.field_validator("filepath", mode="after")
    def ensure_endswith_xlsx(cls, path: pathlib.Path) -> pathlib.Path:
        if path.suffix != ".xlsx":
            raise ValueError("path must be a valid .xlsx file")
        return path.resolve()

    @pydantic.field_validator("filepath_suffix", mode="after")
    def ensure_valid_datetime_format(cls, value: Optional[str]) -> Optional[str]:
        if value is None:
            return value

        try:
            dt.datetime.now(tz=dt.timezone.utc).strftime(value)
        except ValueError:
            raise ValueError("Invalid datetime format for filepath_suffix") from None

        return value

    def __init__(self, **data):
        super().__init__(**data)
        self._workbook = self._load_workbook()

    def _load_workbook(self) -> openpyxl.Workbook:
        try:
            wb = openpyxl.load_workbook(self.make_filename())
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

        return wb

    def make_filename(self) -> pathlib.Path:
        """Enforce the SUFFIX + XLSX extension."""
        if self.filepath_suffix is not None:
            suffix = f"__{dt.datetime.now(tz=dt.timezone.utc).strftime(self.filepath_suffix)}"
            suffix = suffix.replace(":", "-").replace("/", "")
        else:
            suffix = ""

        return self.filepath.parent / f"{self.filepath.stem}{suffix}.xlsx"

    def tab(self, tab_name: str) -> openpyxl.worksheet.worksheet.Worksheet:
        """Fetch the tab. If it does not yet exist, create it."""
        try:
            tab = self._workbook[tab_name]
        except KeyError:
            tab = self._workbook.create_sheet(tab_name)

        return tab

    def __repr__(self):
        return f"<ExcelSyncer path='{self.make_filename()}' in '{self.save_strategy}' mode>"

    # MANDATORY PROTOCOL MEMBERS

    def load(self, tab_name: str) -> _types.TableRowsFormat:
        """Read rows from a tab in the Workbook."""
        tab = self.tab(tab_name)

        if tab.calculate_dimension() == "A1:A1":
            log.warning(f"No data found in tab '{tab_name}'")
            return []

        header = next(tab.iter_rows(min_row=1, max_row=1, values_only=True))
        data = [dict(zip(header, row)) for row in tab.iter_rows(min_row=2, values_only=True)]

        if not data:
            log.warning(f"No data found in tab '{tab_name}'")

        return data

    def dump(self, tab_name: str, *, data: _types.TableRowsFormat) -> None:
        """Write rows to a tab in the Workbook."""
        tab = self.tab(tab_name)

        if not data:
            log.warning(f"No data to write to syncer {self.protocol}.{tab_name}")
            return

        if self.save_strategy == "OVERWRITE":
            # idx = 1 means we should delete the header as well, mostly so we can ensure
            # nothing weird happens here with data/table quality.
            tab.delete_rows(idx=1, amount=tab.max_row + 1)

            # HEADER
            tab.append(list(data[0].keys()))

        # DATA
        for row in data:
            row = sync_utils.format_datetime_values(row, dt_format=self.date_time_format)
            tab.append(list(row.values()))

        self._workbook.save(self.make_filename())
